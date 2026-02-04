import os
import qrcode
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def generate_qr_excel(input_file, output_file):
    # Temporary directory for QR images
    temp_qr_dir = "QRs_Generados"
    if not os.path.exists(temp_qr_dir):
        os.makedirs(temp_qr_dir)

    print(f"Loading {input_file}...")
    try:
        wb = load_workbook(input_file)
        ws = wb.active
    except FileNotFoundError:
        print(f"Error: File {input_file} not found.")
        return

    # Determine the column for QR codes (append to the end)
    qr_col_idx = ws.max_column + 1
    qr_col_letter = get_column_letter(qr_col_idx)

    # Add header
    ws.cell(row=1, column=qr_col_idx, value="QR Code")
    ws.column_dimensions[qr_col_letter].width = 18  # Set column width

    print("Generating QR codes...")
    # Iterate through rows (assuming row 1 is header)
    # enumerate starts at 2 to match Excel row numbers (1-based, skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Create a formatted string of the row data
        data_parts = []
        # Get headers to make data more readable?
        # For now, just dumping values as requested "informacion de toda la fila"
        for cell in row:
            if cell is not None:
                data_parts.append(str(cell))

        row_data = "\n".join(data_parts)

        if not row_data.strip():
            continue  # Skip empty rows

        # Generate QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=2,
        )
        qr.add_data(row_data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Save image
        img_path = os.path.join(temp_qr_dir, f"qr_row_{row_idx}.png")
        img.save(img_path)

        # Insert into Excel
        # openpyxl Image
        img_excel = Image(img_path)
        img_excel.width = 100
        img_excel.height = 100

        # Calculate anchor (cell address)
        anchor = f"{qr_col_letter}{row_idx}"
        ws.add_image(img_excel, anchor)

        # Adjust row height to fit the image
        # Height is in points. 100px is roughly 75 points.
        ws.row_dimensions[row_idx].height = 80

    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print("Done!")


if __name__ == "__main__":
    input_xlsx = "Breidy habilitados Warnes II.xlsx"
    output_xlsx = "Breidy habilitados Warnes II_con_QR.xlsx"

    if os.path.exists(input_xlsx):
        generate_qr_excel(input_xlsx, output_xlsx)
    else:
        print(f"Input file not found: {input_xlsx}")
